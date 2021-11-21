import os
import numpy as np
import pandas as pd
from openpyxl import load_workbook, Workbook

from heat_main import main as heat

def main():
    n = 30
    data_path = './data/'
    output_path = './outputs/'
    output_file = 'data.xlsx'
    programs = ['python', 'cython']
    bottles = ['bottle', 'bottle_medium', 'bottle_large']
    timesteps = [100, 200, 400, 800, 1000]
    data = np.zeros((n * len(bottles), len(timesteps)+1))

    if os.path.exists(output_path+output_file):
        book = load_workbook(output_path+output_file)
    else:
        book = Workbook()
        book.save(output_path+output_file)

    try:
        writer = pd.ExcelWriter(output_path+output_file, engine = 'openpyxl')
        writer.book = book

        for program in programs:
            print(f"Program using {program} start...")
            for k, bottle in enumerate(bottles):
                print(f"\tFor bottle {bottle}...")
                for j, timestep in enumerate(timesteps, start=1):
                    for i in range(n):
                        data[i+k*n, j] = heat(data_path+bottle+'.dat', timesteps=timestep, program=program)
                    print(f"\t\tTimestep {timestep} done.")

                print(f"\tFor bottle {bottle} done.")

            df = pd.DataFrame(data=data, columns=['bottle']+timesteps)
            for k, bottle in enumerate(bottles):
                df.loc[k*n:k*n+n, 'bottle'] = bottle

            means = df.groupby('bottle').mean()

            df.to_excel(writer, sheet_name=program+"-values")
            means.to_excel(writer, sheet_name=program+"-results")
            print(f"Program using {program} done.")

        writer.save()
        writer.close()

    except Exception as e:
        print(e)


if __name__ == '__main__':
    main()
